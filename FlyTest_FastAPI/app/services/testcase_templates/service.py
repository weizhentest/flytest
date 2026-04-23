from __future__ import annotations

import io
import os
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.core.errors import AppError
from app.db.models.auth import User
from app.db.models.testcase_templates import ImportExportTemplate
from app.repositories.testcase_templates import TestcaseTemplateRepository

REPO_ROOT = Path(__file__).resolve().parents[4]
DJANGO_MEDIA_ROOT = REPO_ROOT / "FlyTest_Django" / "media"


TEMPLATE_TYPE_LABELS = {
    "import": "瀵煎叆",
    "export": "瀵煎嚭",
    "both": "瀵煎叆/瀵煎嚭",
}

STEP_PARSING_MODE_LABELS = {
    "single_cell": "鍗曞厓鏍煎悎骞舵楠?",
    "multi_row": "澶氳姝ラ",
}


def _template_file_url(value: str | None) -> str | None:
    if not value:
        return None
    normalized = str(value).replace("\\", "/").lstrip("/")
    return f"/media/{normalized}"


def _serialize_template(template: ImportExportTemplate) -> dict:
    return {
        "id": template.id,
        "name": template.name,
        "template_type": template.template_type,
        "template_type_display": TEMPLATE_TYPE_LABELS.get(template.template_type, template.template_type),
        "description": template.description,
        "sheet_name": template.sheet_name,
        "template_file": _template_file_url(template.template_file),
        "template_headers": list(template.template_headers or []),
        "header_row": template.header_row,
        "data_start_row": template.data_start_row,
        "field_mappings": dict(template.field_mappings or {}),
        "value_transformations": dict(template.value_transformations or {}),
        "step_parsing_mode": template.step_parsing_mode,
        "step_parsing_mode_display": STEP_PARSING_MODE_LABELS.get(template.step_parsing_mode, template.step_parsing_mode),
        "step_config": dict(template.step_config or {}),
        "module_path_delimiter": template.module_path_delimiter,
        "is_active": bool(template.is_active),
        "creator": template.creator_id,
        "creator_name": template.creator.username if template.creator else None,
        "created_at": template.created_at.isoformat() if template.created_at else "",
        "updated_at": template.updated_at.isoformat() if template.updated_at else "",
    }


def _serialize_template_list_item(template: ImportExportTemplate) -> dict:
    return {
        "id": template.id,
        "name": template.name,
        "template_type": template.template_type,
        "template_type_display": TEMPLATE_TYPE_LABELS.get(template.template_type, template.template_type),
        "description": template.description,
        "is_active": bool(template.is_active),
        "creator_name": template.creator.username if template.creator else None,
        "created_at": template.created_at.isoformat() if template.created_at else "",
        "updated_at": template.updated_at.isoformat() if template.updated_at else "",
    }


def list_templates(*, db: Session, template_type: str | None = None, is_active: bool | None = None) -> list[dict]:
    queryset = TestcaseTemplateRepository(db).list_templates(template_type=template_type, is_active=is_active)
    return [_serialize_template_list_item(item) for item in queryset]


def get_template(*, db: Session, template_id: int) -> ImportExportTemplate:
    template = TestcaseTemplateRepository(db).get_template(template_id=template_id)
    if not template:
        raise AppError("妯℃澘涓嶅瓨鍦?", status_code=404)
    return template


def get_template_detail(*, db: Session, template_id: int) -> dict:
    return _serialize_template(get_template(db=db, template_id=template_id))


def create_template(*, db: Session, payload: dict, creator: User) -> dict:
    repo = TestcaseTemplateRepository(db)
    name = str(payload.get("name") or "").strip()
    if not name:
        raise AppError("妯℃澘鍚嶇О涓嶈兘涓虹┖", status_code=400)
    if repo.get_by_name(name=name):
        raise AppError("妯℃澘鍚嶇О宸插瓨鍦?", status_code=400)

    template = repo.create_template(
        ImportExportTemplate(
            name=name,
            template_type=str(payload.get("template_type") or "import"),
            description=payload.get("description"),
            sheet_name=payload.get("sheet_name"),
            template_headers=list(payload.get("template_headers") or []),
            header_row=int(payload.get("header_row") or 1),
            data_start_row=int(payload.get("data_start_row") or 2),
            field_mappings=dict(payload.get("field_mappings") or {}),
            value_transformations=dict(payload.get("value_transformations") or {}),
            step_parsing_mode=str(payload.get("step_parsing_mode") or "single_cell"),
            step_config=dict(payload.get("step_config") or {}),
            module_path_delimiter=str(payload.get("module_path_delimiter") or "/"),
            is_active=bool(payload.get("is_active", True)),
            creator_id=creator.id,
            created_at=datetime.now(),
            updated_at=datetime.now(),
        )
    )
    db.commit()
    template = repo.get_template(template_id=template.id)
    return _serialize_template(template)


def update_template(*, db: Session, template_id: int, payload: dict) -> dict:
    template = get_template(db=db, template_id=template_id)
    repo = TestcaseTemplateRepository(db)
    if "name" in payload and payload.get("name") is not None:
        name = str(payload.get("name") or "").strip()
        if not name:
            raise AppError("妯℃澘鍚嶇О涓嶈兘涓虹┖", status_code=400)
        existing = repo.get_by_name(name=name)
        if existing and existing.id != template.id:
            raise AppError("妯℃澘鍚嶇О宸插瓨鍦?", status_code=400)
        template.name = name
    for field in ["description", "sheet_name", "field_mappings", "value_transformations", "step_config"]:
        if field in payload:
            setattr(template, field, payload.get(field))
    if "template_type" in payload:
        template.template_type = str(payload.get("template_type") or template.template_type)
    if "template_headers" in payload:
        template.template_headers = list(payload.get("template_headers") or [])
    if "header_row" in payload:
        template.header_row = int(payload.get("header_row") or template.header_row)
    if "data_start_row" in payload:
        template.data_start_row = int(payload.get("data_start_row") or template.data_start_row)
    if "step_parsing_mode" in payload:
        template.step_parsing_mode = str(payload.get("step_parsing_mode") or template.step_parsing_mode)
    if "module_path_delimiter" in payload:
        template.module_path_delimiter = str(payload.get("module_path_delimiter") or template.module_path_delimiter)
    if "is_active" in payload:
        template.is_active = bool(payload.get("is_active"))
    template.updated_at = datetime.now()
    db.add(template)
    db.commit()
    db.refresh(template)
    return _serialize_template(template)


def delete_template(*, db: Session, template_id: int) -> None:
    template = get_template(db=db, template_id=template_id)
    template_file = template.template_file
    db.delete(template)
    db.commit()
    if template_file:
        media_file = os.path.join(str((__import__("pathlib").Path(__file__).resolve().parents[4] / "FlyTest_Django" / "media")), template_file)
        if os.path.exists(media_file):
            try:
                os.remove(media_file)
            except Exception:
                pass


def parse_excel_headers(*, file_bytes: bytes, sheet_name: str = "", header_row: int = 1) -> dict:
    try:
        workbook = load_workbook(filename=io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001
        raise AppError(f"瑙ｆ瀽 Excel 鏂囦欢澶辫触: {exc}", status_code=400)

    sheet_names = workbook.sheetnames
    worksheet = workbook[sheet_name] if sheet_name and sheet_name in sheet_names else workbook.active

    headers: list[str] = []
    for cell in worksheet[header_row]:
        headers.append(str(cell.value).strip() if cell.value is not None else "")
    while headers and headers[-1] == "":
        headers.pop()

    sample_data = []
    for row in worksheet.iter_rows(min_row=header_row + 1, max_row=header_row + 3):
        row_data = {}
        for col_idx, cell in enumerate(row):
            if col_idx < len(headers) and headers[col_idx]:
                row_data[headers[col_idx]] = str(cell.value) if cell.value is not None else ""
        if any(row_data.values()):
            sample_data.append(row_data)

    row_count = (worksheet.max_row or 0) - header_row
    workbook.close()
    return {
        "headers": headers,
        "sheet_names": sheet_names,
        "row_count": max(row_count, 0),
        "sample_data": sample_data,
    }


def get_field_options() -> dict:
    return {
        "fields": [
            {"value": "name", "label": "鐢ㄤ緥鍚嶇О", "required": True},
            {"value": "module", "label": "鎵€灞炴ā鍧?", "required": True},
            {"value": "precondition", "label": "鍓嶇疆鏉′欢", "required": False},
            {"value": "level", "label": "鐢ㄤ緥绛夌骇", "required": False, "has_transform": True},
            {"value": "notes", "label": "澶囨敞", "required": False},
            {"value": "steps", "label": "姝ラ鎻忚堪", "required": False, "is_step_field": True},
            {"value": "expected_results", "label": "棰勬湡缁撴灉", "required": False, "is_step_field": True},
        ],
        "level_options": [
            {"value": "P0", "label": "P0 - 鏈€楂樹紭鍏堢骇"},
            {"value": "P1", "label": "P1 - 楂樹紭鍏堢骇"},
            {"value": "P2", "label": "P2 - 涓紭鍏堢骇"},
            {"value": "P3", "label": "P3 - 浣庝紭鍏堢骇"},
        ],
    }


def duplicate_template(*, db: Session, template_id: int, creator: User) -> dict:
    template = get_template(db=db, template_id=template_id)
    repo = TestcaseTemplateRepository(db)
    new_name = f"{template.name} (鍓湰)"
    counter = 1
    while repo.get_by_name(name=new_name):
        counter += 1
        new_name = f"{template.name} (鍓湰{counter})"

    duplicated = repo.create_template(
        ImportExportTemplate(
            name=new_name,
            template_type=template.template_type,
            description=template.description,
            sheet_name=template.sheet_name,
            template_file=template.template_file,
            template_headers=list(template.template_headers or []),
            header_row=template.header_row,
            data_start_row=template.data_start_row,
            field_mappings=dict(template.field_mappings or {}),
            value_transformations=dict(template.value_transformations or {}),
            step_parsing_mode=template.step_parsing_mode,
            step_config=dict(template.step_config or {}),
            module_path_delimiter=template.module_path_delimiter,
            is_active=True,
            creator_id=creator.id,
            created_at=datetime.now(),
            updated_at=datetime.now(),
        )
    )
    db.commit()
    duplicated = repo.get_template(template_id=duplicated.id)
    return _serialize_template(duplicated)


def upload_template_file(*, db: Session, template_id: int, filename: str, file_bytes: bytes) -> dict:
    template = get_template(db=db, template_id=template_id)
    if not file_bytes:
        raise AppError("Uploaded file is empty", status_code=400)

    target_dir = DJANGO_MEDIA_ROOT / "testcase_templates"
    target_dir.mkdir(parents=True, exist_ok=True)
    stored_name = filename or "template.xlsx"
    relative_path = Path("testcase_templates") / stored_name
    full_path = DJANGO_MEDIA_ROOT / relative_path

    try:
        with open(full_path, "wb") as handle:
            handle.write(file_bytes)

        workbook = load_workbook(filename=io.BytesIO(file_bytes), read_only=True, data_only=True)
        sheet_names = workbook.sheetnames

        if template.sheet_name and template.sheet_name in sheet_names:
            worksheet = workbook[template.sheet_name]
        else:
            worksheet = workbook.active
            if not template.sheet_name:
                template.sheet_name = worksheet.title

        header_row = template.header_row or 1
        headers: list[str] = []
        for cell in worksheet[header_row]:
            headers.append(str(cell.value).strip() if cell.value is not None else "")
        while headers and headers[-1] == "":
            headers.pop()

        workbook.close()

        template.template_file = relative_path.as_posix()
        template.template_headers = headers
        template.updated_at = datetime.now()
        db.add(template)
        db.commit()
        db.refresh(template)

        return {
            "template_id": template.id,
            "template_file": _template_file_url(template.template_file),
            "template_headers": headers,
            "sheet_names": sheet_names,
        }
    except AppError:
        raise
    except Exception as exc:  # noqa: BLE001
        raise AppError(f"Upload template file failed: {exc}", status_code=400)
