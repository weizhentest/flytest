from rest_framework import viewsets, status, permissions
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser
from django.db import transaction
from django.core.files.base import ContentFile
from openpyxl import load_workbook
import io

from .models import ImportExportTemplate
from .serializers import (
    ImportExportTemplateSerializer,
    ImportExportTemplateListSerializer,
    ParseHeadersRequestSerializer,
    ParseHeadersResponseSerializer,
)


class ImportExportTemplateViewSet(viewsets.ModelViewSet):
    """
    导入导出模版管理 ViewSet

    提供模版的 CRUD 操作，以及解析 Excel 表头等辅助功能
    """
    queryset = ImportExportTemplate.objects.all()
    permission_classes = [permissions.IsAuthenticated]

    def get_serializer_class(self):
        if self.action == 'list':
            return ImportExportTemplateListSerializer
        return ImportExportTemplateSerializer

    def get_queryset(self):
        """支持按类型和状态筛选"""
        queryset = ImportExportTemplate.objects.all()

        # 按模版类型筛选
        template_type = self.request.query_params.get('template_type')
        if template_type:
            queryset = queryset.filter(template_type=template_type)

        # 按启用状态筛选
        is_active = self.request.query_params.get('is_active')
        if is_active is not None:
            queryset = queryset.filter(is_active=is_active.lower() == 'true')

        return queryset

    @action(detail=False, methods=['post'], parser_classes=[MultiPartParser, FormParser])
    def parse_headers(self, request):
        """
        解析上传的 Excel 文件表头

        上传 Excel 文件，返回表头列名列表，用于配置字段映射
        """
        serializer = ParseHeadersRequestSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        file = serializer.validated_data['file']
        sheet_name = serializer.validated_data.get('sheet_name', '')
        header_row = serializer.validated_data.get('header_row', 1)

        try:
            # 读取 Excel 文件
            workbook = load_workbook(filename=io.BytesIO(file.read()), read_only=True, data_only=True)

            # 获取所有工作表名称
            sheet_names = workbook.sheetnames

            # 选择工作表
            if sheet_name and sheet_name in sheet_names:
                worksheet = workbook[sheet_name]
            else:
                worksheet = workbook.active

            # 读取表头行
            headers = []
            for cell in worksheet[header_row]:
                if cell.value is not None:
                    headers.append(str(cell.value).strip())
                else:
                    headers.append('')

            # 去除尾部空列
            while headers and headers[-1] == '':
                headers.pop()

            # 读取样例数据（表头后的前3行）
            sample_data = []
            max_sample_rows = 3
            for row_idx, row in enumerate(worksheet.iter_rows(min_row=header_row + 1, max_row=header_row + max_sample_rows), start=1):
                row_data = {}
                for col_idx, cell in enumerate(row):
                    if col_idx < len(headers) and headers[col_idx]:
                        row_data[headers[col_idx]] = str(cell.value) if cell.value is not None else ''
                if any(row_data.values()):  # 跳过空行
                    sample_data.append(row_data)

            # 计算数据行数
            row_count = worksheet.max_row - header_row if worksheet.max_row else 0

            workbook.close()

            response_data = {
                'headers': headers,
                'sheet_names': sheet_names,
                'row_count': row_count,
                'sample_data': sample_data,
            }

            return Response(response_data, status=status.HTTP_200_OK)

        except Exception as e:
            return Response(
                {'error': f'解析 Excel 文件失败: {str(e)}'},
                status=status.HTTP_400_BAD_REQUEST
            )

    @action(detail=True, methods=['post'], parser_classes=[MultiPartParser, FormParser], url_path='upload-template-file')
    def upload_template_file(self, request, pk=None):
        """
        上传并绑定模板 Excel 文件

        将用户上传的 Excel 文件保存到模板上，导出时可基于该文件保留标题行/合并单元格/样式等结构。
        同时会按模板的 header_row/sheet_name 重新解析并更新 template_headers。
        """
        template = self.get_object()
        file = request.data.get('file')
        if not file:
            return Response({'error': '请上传 Excel 文件'}, status=status.HTTP_400_BAD_REQUEST)

        file_bytes = file.read()
        if not file_bytes:
            return Response({'error': '上传的文件为空'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            # 保存模板文件
            template.template_file.save(file.name, ContentFile(file_bytes), save=False)

            # 解析表头并更新 template_headers（保持列顺序，保留中间空列，仅去除尾部空列）
            workbook = load_workbook(filename=io.BytesIO(file_bytes), read_only=True, data_only=True)
            sheet_names = workbook.sheetnames

            if template.sheet_name and template.sheet_name in sheet_names:
                worksheet = workbook[template.sheet_name]
            else:
                worksheet = workbook.active
                if not template.sheet_name:
                    template.sheet_name = worksheet.title

            header_row = template.header_row or 1
            headers = []
            for cell in worksheet[header_row]:
                if cell.value is not None:
                    headers.append(str(cell.value).strip())
                else:
                    headers.append('')
            while headers and headers[-1] == '':
                headers.pop()

            workbook.close()

            template.template_headers = headers
            template.save(update_fields=['template_file', 'template_headers', 'sheet_name', 'updated_at'])

            return Response(
                {
                    'template_id': template.id,
                    'template_file': getattr(template.template_file, 'url', None),
                    'template_headers': headers,
                    'sheet_names': sheet_names,
                },
                status=status.HTTP_200_OK
            )
        except Exception as e:
            return Response(
                {'error': f'上传并解析模板文件失败: {str(e)}'},
                status=status.HTTP_400_BAD_REQUEST
            )

    @action(detail=False, methods=['get'])
    def field_options(self, request):
        """
        获取可映射的 TestCase 字段选项

        返回用例模型的可映射字段列表，用于前端字段映射配置
        """
        field_options = [
            {'value': 'name', 'label': '用例名称', 'required': True},
            {'value': 'module', 'label': '所属模块', 'required': True},
            {'value': 'precondition', 'label': '前置条件', 'required': False},
            {'value': 'level', 'label': '用例等级', 'required': False, 'has_transform': True},
            {'value': 'notes', 'label': '备注', 'required': False},
            {'value': 'steps', 'label': '步骤描述', 'required': False, 'is_step_field': True},
            {'value': 'expected_results', 'label': '预期结果', 'required': False, 'is_step_field': True},
        ]

        level_options = [
            {'value': 'P0', 'label': 'P0 - 最高优先级'},
            {'value': 'P1', 'label': 'P1 - 高优先级'},
            {'value': 'P2', 'label': 'P2 - 中优先级'},
            {'value': 'P3', 'label': 'P3 - 低优先级'},
        ]

        return Response({
            'fields': field_options,
            'level_options': level_options,
        })

    @action(detail=True, methods=['post'])
    def duplicate(self, request, pk=None):
        """
        复制模版

        创建一个现有模版的副本，名称后缀添加 "(副本)"
        """
        template = self.get_object()

        # 生成新名称
        new_name = f"{template.name} (副本)"
        counter = 1
        while ImportExportTemplate.objects.filter(name=new_name).exists():
            counter += 1
            new_name = f"{template.name} (副本{counter})"

        # 创建副本
        new_template = ImportExportTemplate.objects.create(
            name=new_name,
            template_type=template.template_type,
            description=template.description,
            sheet_name=template.sheet_name,
            header_row=template.header_row,
            data_start_row=template.data_start_row,
            field_mappings=template.field_mappings.copy() if template.field_mappings else {},
            value_transformations=template.value_transformations.copy() if template.value_transformations else {},
            step_parsing_mode=template.step_parsing_mode,
            step_config=template.step_config.copy() if template.step_config else {},
            module_path_delimiter=template.module_path_delimiter,
            is_active=True,
            creator=request.user,
        )

        serializer = ImportExportTemplateSerializer(new_template, context={'request': request})
        return Response(serializer.data, status=status.HTTP_201_CREATED)
