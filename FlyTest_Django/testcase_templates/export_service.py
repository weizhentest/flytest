"""
用例导出服务 - 支持模版化导出
"""

import io
import posixpath
import re
import zipfile
import xml.etree.ElementTree as ET
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment


class TestCaseExportService:
    """测试用例导出服务"""


    # 内部字段名到默认列名的映射
    DEFAULT_FIELD_NAMES = {
        "name": "用例名称",
        "module": "所属模块",
        "precondition": "前置条件",
        "level": "用例等级",
        "notes": "备注",
        "steps": "步骤描述",
        "expected_results": "预期结果",
    }

    def __init__(self, template=None):
        """
        初始化导出服务
        :param template: ImportExportTemplate 实例，为 None 时使用默认格式
        """
        self.template = template

    def export(self, queryset, project_name: str) -> tuple:
        """
        导出用例到 Excel
        :param queryset: TestCase QuerySet
        :param project_name: 项目名称（用于文件名）
        :return: (bytes, filename)
        """
        # 获取字段映射（模版或默认）
        field_mappings = self._get_field_mappings()
        value_transformations = self._get_value_transformations()

        using_template_file = bool(
            self.template
            and getattr(self.template, "template_file", None)
            and self.template.template_file
        )
        if using_template_file:
            excel_bytes = self._export_using_template_file_bytes(
                queryset, field_mappings, value_transformations
            )
        else:
            excel_bytes = self._export_default_bytes(
                queryset, field_mappings, value_transformations
            )

        filename = f"{project_name}_测试用例.xlsx"
        return excel_bytes, filename

    def _export_default_bytes(
        self, queryset, field_mappings: dict, value_transformations: dict
    ) -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = (
            self.template.sheet_name
            if self.template and self.template.sheet_name
            else "测试用例"
        )

        # 构建表头
        headers = self._get_headers(field_mappings)
        header_row = self.template.header_row if self.template else 1
        header_to_col = {}
        for col, header in enumerate(headers, 1):
            cell_value = header if header is not None else ""
            cell = ws.cell(row=header_row, column=col, value=cell_value)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
            if cell_value:
                header_to_col[str(cell_value).strip()] = col

        # 写入数据
        data_start_row = self.template.data_start_row if self.template else 2
        for row_idx, testcase in enumerate(queryset, data_start_row):
            self._write_testcase_row(
                ws,
                row_idx,
                testcase,
                field_mappings,
                value_transformations,
                header_to_col=header_to_col if header_to_col else None,
            )

        # 调整列宽
        for col in range(1, len(headers) + 1):
            col_letter = self._get_column_letter(col)
            ws.column_dimensions[col_letter].width = 20

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    def _export_using_template_file_bytes(
        self, queryset, field_mappings: dict, value_transformations: dict
    ) -> bytes:
        """
        基于用户上传的模板文件直接打补丁写入数据，尽量做到“模板原样不变，只新增数据”。

        说明：openpyxl 在“读入并另存”时无法完整保留某些 Excel 特性（如单元格内局部红字的 rich text runs），
        因此这里采用 xlsx(zip+xml) 方式仅修改目标数据单元格，避免破坏模板里的富文本/样式/合并等。
        """
        self.template.template_file.open("rb")
        try:
            template_bytes = self.template.template_file.read()
        finally:
            try:
                self.template.template_file.close()
            except Exception:
                pass

        header_row = self.template.header_row if self.template else 1
        data_start_row = self.template.data_start_row if self.template else 2
        data_count = self._safe_count(queryset)
        need_last_row = (
            data_start_row + data_count - 1 if data_count > 0 else data_start_row
        )

        patcher = _XlsxTemplatePatcher(
            xlsx_bytes=template_bytes,
            sheet_name=(self.template.sheet_name or None),
            header_row=header_row,
            data_start_row=data_start_row,
            need_last_row=need_last_row,
        )

        header_to_col = patcher.read_header_to_col()

        # 写入数据（按表头列定位写入；未映射的列保持模板原样）
        row_idx = data_start_row
        for testcase in queryset:
            values_by_header = {}
            for field in [
                "name",
                "module",
                "precondition",
                "steps",
                "expected_results",
                "level",
                "notes",
            ]:
                if field not in field_mappings:
                    continue
                header_name = field_mappings.get(field)
                if not header_name:
                    continue
                col_num = header_to_col.get(str(header_name).strip())
                if not col_num:
                    continue
                values_by_header[col_num] = self._get_field_value(
                    testcase, field, value_transformations
                )
            patcher.write_row_values(row_idx=row_idx, values_by_col=values_by_header)
            row_idx += 1

        patcher.extend_table_and_filter_refs()
        patcher.update_dimension()
        return patcher.to_bytes()

    def _safe_count(self, queryset) -> int:
        try:
            return int(queryset.count())
        except Exception:
            try:
                return len(queryset)
            except Exception:
                return 0

    def _get_field_mappings(self) -> dict:
        """获取字段映射配置"""
        if self.template and self.template.field_mappings:
            return self.template.field_mappings
        # 模版存在但未配置字段映射时，仍使用默认字段名尝试填充（尤其是导出模版场景）
        if self.template:
            return self.DEFAULT_FIELD_NAMES
        return self.DEFAULT_FIELD_NAMES

    def _get_value_transformations(self) -> dict:
        """获取值转换规则（反向映射，用于导出）"""
        if not self.template or not self.template.value_transformations:
            return {}

        # 构建反向映射：内部值 -> 显示值
        reverse_transformations = {}
        for field, mapping in self.template.value_transformations.items():
            reverse_transformations[field] = {v: k for k, v in mapping.items()}
        return reverse_transformations

    def _get_headers(self, field_mappings: dict) -> list:
        """
        获取导出表头列表

        优先使用用户上传/解析得到的 template_headers，以保持模版列顺序；
        若不存在则按固定字段顺序生成。
        """
        if self.template and getattr(self.template, "template_headers", None):
            headers = list(self.template.template_headers or [])
            # 若字段映射中包含未出现在 template_headers 的列名，追加到末尾避免丢列
            for col_name in (field_mappings or {}).values():
                if col_name and col_name not in headers:
                    headers.append(col_name)
            return headers
        return self._build_headers_by_field_order(field_mappings)

    def _build_headers_by_field_order(self, field_mappings: dict) -> list:
        """构建表头列表"""
        # 按照固定顺序构建表头
        field_order = [
            "name",
            "module",
            "precondition",
            "steps",
            "expected_results",
            "level",
            "notes",
        ]
        headers = []
        for field in field_order:
            if field in field_mappings:
                headers.append(field_mappings[field])
        return headers

    def _write_testcase_row(
        self,
        ws,
        row: int,
        testcase,
        field_mappings: dict,
        value_transformations: dict,
        header_to_col: dict | None = None,
    ):
        """写入单个用例行"""
        field_order = [
            "name",
            "module",
            "precondition",
            "steps",
            "expected_results",
            "level",
            "notes",
        ]
        col = 1

        for field in field_order:
            if field not in field_mappings:
                continue

            value = self._get_field_value(testcase, field, value_transformations)
            if header_to_col is not None:
                header_name = field_mappings.get(field)
                if not header_name:
                    continue
                col = header_to_col.get(str(header_name).strip())
                if not col:
                    continue
                ws.cell(row=row, column=col, value=value)
            else:
                # 兼容旧逻辑：按字段顺序顺次写入
                ws.cell(row=row, column=col, value=value)
                col += 1

    def _get_field_value(
        self, testcase, field: str, value_transformations: dict
    ) -> str:
        """获取字段值"""
        if field == "name":
            return testcase.name
        elif field == "module":
            return self._get_module_path(testcase.module)
        elif field == "precondition":
            return testcase.precondition or ""
        elif field == "level":
            value = testcase.level
            # 应用反向值转换
            if (
                "level" in value_transformations
                and value in value_transformations["level"]
            ):
                return value_transformations["level"][value]
            return value
        elif field == "notes":
            return testcase.notes or ""
        elif field == "steps":
            return self._format_steps_desc(testcase.steps.all())
        elif field == "expected_results":
            return self._format_expected_results(testcase.steps.all())
        return ""

    def _get_module_path(self, module) -> str:
        """获取模块完整路径"""
        if not module:
            return ""

        delimiter = self.template.module_path_delimiter if self.template else "/"
        path_parts = []
        current = module
        while current:
            path_parts.insert(0, current.name)
            current = current.parent

        return delimiter + delimiter.join(path_parts)

    def _format_steps_desc(self, steps) -> str:
        """格式化步骤描述"""
        step_list = []
        for step in steps.order_by("step_number"):
            step_list.append(f"[{step.step_number}]{step.description}")
        return "\n".join(step_list)

    def _format_expected_results(self, steps) -> str:
        """格式化预期结果"""
        result_list = []
        for step in steps.order_by("step_number"):
            result_list.append(f"[{step.step_number}]{step.expected_result}")
        return "\n".join(result_list)

    def _get_column_letter(self, col_idx: int) -> str:
        """将列索引转换为列字母"""
        result = ""
        while col_idx > 0:
            col_idx, remainder = divmod(col_idx - 1, 26)
            result = chr(65 + remainder) + result
        return result


class _XlsxTemplatePatcher:
    """
    仅修改目标 sheet 的数据单元格，尽量不破坏模板中的富文本/样式/合并等。
    """

    NS_MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    # 用于 workbook/sheet XML 中的 r:id 等属性
    NS_REL = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
    # 用于 *.rels 文件中的 Relationship 元素
    NS_PKG_REL = "http://schemas.openxmlformats.org/package/2006/relationships"
    XML_SPACE = "{http://www.w3.org/XML/1998/namespace}space"

    def __init__(
        self,
        xlsx_bytes: bytes,
        sheet_name: str | None,
        header_row: int,
        data_start_row: int,
        need_last_row: int,
    ):
        self._zip_in = zipfile.ZipFile(io.BytesIO(xlsx_bytes), "r")
        self._original_entries = {
            zi.filename: self._zip_in.read(zi.filename)
            for zi in self._zip_in.infolist()
        }
        self._sheet_name = sheet_name
        self._header_row = int(header_row)
        self._data_start_row = int(data_start_row)
        self._need_last_row = int(need_last_row)

        ET.register_namespace("", self.NS_MAIN)
        ET.register_namespace("r", self.NS_REL)

        self._shared_strings = None
        self._sheet_path = self._resolve_sheet_path()
        self._sheet_xml = self._parse_xml(self._original_entries[self._sheet_path])
        self._sheet_rels_path = self._guess_sheet_rels_path(self._sheet_path)
        self._sheet_rels_xml = (
            self._parse_xml(self._original_entries[self._sheet_rels_path])
            if self._sheet_rels_path in self._original_entries
            else None
        )

        self._merged_ranges = self._parse_merged_ranges()
        self._style_template = self._capture_style_template()

    def _q(self, tag: str) -> str:
        return f"{{{self.NS_MAIN}}}{tag}"

    def _qr(self, tag: str) -> str:
        return f"{{{self.NS_REL}}}{tag}"

    def _parse_xml(self, xml_bytes: bytes) -> ET.ElementTree:
        return ET.ElementTree(ET.fromstring(xml_bytes))

    def _resolve_sheet_path(self) -> str:
        wb_xml = self._parse_xml(self._original_entries["xl/workbook.xml"])
        wb_rels = self._parse_xml(self._original_entries["xl/_rels/workbook.xml.rels"])

        sheets_el = wb_xml.getroot().find(self._q("sheets"))
        if sheets_el is None:
            raise ValueError("workbook.xml 缺少 sheets")

        sheet_elems = list(sheets_el.findall(self._q("sheet")))
        if not sheet_elems:
            raise ValueError("workbook.xml 未找到任何 sheet")

        chosen = None
        if self._sheet_name:
            for s in sheet_elems:
                if s.attrib.get("name") == self._sheet_name:
                    chosen = s
                    break
        if chosen is None:
            # activeTab 优先，否则选第一个
            active_tab = 0
            view = wb_xml.getroot().find(
                f"{self._q('bookViews')}/{self._q('workbookView')}"
            )
            if view is not None and view.attrib.get("activeTab") is not None:
                try:
                    active_tab = int(view.attrib.get("activeTab") or 0)
                except ValueError:
                    active_tab = 0
            if 0 <= active_tab < len(sheet_elems):
                chosen = sheet_elems[active_tab]
            else:
                chosen = sheet_elems[0]

        rid = chosen.attrib.get(self._qr("id"))
        if not rid:
            raise ValueError("sheet 缺少 r:id")

        rel = wb_rels.getroot().find(
            f".//{{{self.NS_PKG_REL}}}Relationship[@Id='{rid}']"
        )
        if rel is None:
            raise ValueError(f"workbook.xml.rels 未找到 {rid}")

        target = rel.attrib.get("Target")
        if not target:
            raise ValueError(f"{rid} 缺少 Target")
        # Target 形如 'worksheets/sheet1.xml'
        return posixpath.normpath(posixpath.join("xl", target))

    def _guess_sheet_rels_path(self, sheet_path: str) -> str:
        # 例如：xl/worksheets/sheet1.xml -> xl/worksheets/_rels/sheet1.xml.rels
        base_dir = posixpath.dirname(sheet_path)
        base_name = posixpath.basename(sheet_path)
        return posixpath.join(base_dir, "_rels", f"{base_name}.rels")

    def _load_shared_strings(self):
        if self._shared_strings is not None:
            return
        if "xl/sharedStrings.xml" not in self._original_entries:
            self._shared_strings = []
            return
        sst_xml = self._parse_xml(self._original_entries["xl/sharedStrings.xml"])
        sst_root = sst_xml.getroot()
        strings = []
        for si in sst_root.findall(self._q("si")):
            # si 可包含一个 <t> 或多个 <r><t>（富文本）
            t = si.find(self._q("t"))
            if t is not None:
                strings.append(t.text or "")
                continue
            runs = []
            for r in si.findall(self._q("r")):
                rt = r.find(self._q("t"))
                if rt is not None and rt.text:
                    runs.append(rt.text)
            strings.append("".join(runs))
        self._shared_strings = strings

    def _cell_ref_to_col_row(self, ref: str) -> tuple[int, int]:
        # 支持 A1 / $A$1 / a1 这类引用
        m = re.match(r"^\$?([A-Za-z]+)\$?(\d+)$", ref)
        if not m:
            raise ValueError(f"无效单元格引用: {ref}")
        col_letters, row_str = m.group(1), m.group(2)
        col = 0
        for ch in col_letters.upper():
            col = col * 26 + (ord(ch) - 64)
        return col, int(row_str)

    def _col_to_letters(self, col: int) -> str:
        result = ""
        while col > 0:
            col, rem = divmod(col - 1, 26)
            result = chr(65 + rem) + result
        return result

    def _get_cell_text(self, c: ET.Element) -> str:
        t = c.attrib.get("t")
        if t == "inlineStr":
            is_el = c.find(self._q("is"))
            if is_el is None:
                return ""
            tt = is_el.find(self._q("t"))
            if tt is not None:
                return tt.text or ""
            runs = []
            for r in is_el.findall(self._q("r")):
                rt = r.find(self._q("t"))
                if rt is not None and rt.text:
                    runs.append(rt.text)
            return "".join(runs)
        if t == "s":
            self._load_shared_strings()
            v = c.find(self._q("v"))
            if v is None or v.text is None:
                return ""
            try:
                idx = int(v.text)
            except ValueError:
                return ""
            if 0 <= idx < len(self._shared_strings):
                return self._shared_strings[idx]
            return ""
        v = c.find(self._q("v"))
        return v.text or "" if v is not None else ""

    def _sheet_data(self) -> ET.Element:
        sheet_data = self._sheet_xml.getroot().find(self._q("sheetData"))
        if sheet_data is None:
            sheet_data = ET.SubElement(self._sheet_xml.getroot(), self._q("sheetData"))
        return sheet_data

    def _get_row_el(self, row_idx: int) -> ET.Element | None:
        for row_el in self._sheet_data().findall(self._q("row")):
            if int(row_el.attrib.get("r") or 0) == row_idx:
                return row_el
        return None

    def _parse_merged_ranges(self) -> list[tuple[int, int, int, int]]:
        merged = []
        merged_cells = self._sheet_xml.getroot().find(self._q("mergeCells"))
        if merged_cells is None:
            return merged
        for mc in merged_cells.findall(self._q("mergeCell")):
            ref = mc.attrib.get("ref")
            if not ref or ":" not in ref:
                continue
            a, b = ref.split(":")
            c1, r1 = self._cell_ref_to_col_row(a)
            c2, r2 = self._cell_ref_to_col_row(b)
            merged.append((min(r1, r2), min(c1, c2), max(r1, r2), max(c1, c2)))
        return merged

    def _redirect_merged_target(self, row: int, col: int) -> tuple[int, int]:
        for r1, c1, r2, c2 in self._merged_ranges:
            if r1 <= row <= r2 and c1 <= col <= c2:
                return r1, c1
        return row, col

    def _capture_style_template(self) -> dict:
        """
        捕获数据起始行的行属性与各列 cell 的样式索引(s)，用于在模板行不足时扩展新行。
        """
        style_row = self._get_row_el(self._data_start_row)
        if style_row is None:
            # 退化为最后一行
            all_rows = self._sheet_data().findall(self._q("row"))
            style_row = all_rows[-1] if all_rows else None
        if style_row is None:
            return {"row_attrib": {}, "col_style": {}}

        row_attrib = dict(style_row.attrib)
        row_attrib.pop("r", None)
        col_style = {}
        for c in style_row.findall(self._q("c")):
            ref = c.attrib.get("r")
            if not ref:
                continue
            col, _ = self._cell_ref_to_col_row(ref)
            s = c.attrib.get("s")
            if s is not None:
                col_style[col] = s
        return {"row_attrib": row_attrib, "col_style": col_style}

    def read_header_to_col(self) -> dict[str, int]:
        header_to_col: dict[str, int] = {}
        header_row_el = self._get_row_el(self._header_row)
        if header_row_el is None:
            # 兜底：使用已保存的 template_headers 顺序（无法恢复空列/跳列）
            return header_to_col
        for c in header_row_el.findall(self._q("c")):
            ref = c.attrib.get("r")
            if not ref:
                continue
            col, _ = self._cell_ref_to_col_row(ref)
            text = (self._get_cell_text(c) or "").strip()
            if text:
                header_to_col[text] = col
        return header_to_col

    def _ensure_row_exists(self, row_idx: int) -> ET.Element:
        row_el = self._get_row_el(row_idx)
        if row_el is not None:
            return row_el

        row_el = ET.Element(
            self._q("row"),
            {"r": str(row_idx), **(self._style_template.get("row_attrib") or {})},
        )
        sheet_data = self._sheet_data()
        # 按 r 顺序插入
        inserted = False
        for i, existing in enumerate(sheet_data.findall(self._q("row"))):
            er = int(existing.attrib.get("r") or 0)
            if er > row_idx:
                sheet_data.insert(i, row_el)
                inserted = True
                break
        if not inserted:
            sheet_data.append(row_el)
        return row_el

    def _get_or_create_cell(
        self, row_el: ET.Element, row_idx: int, col: int
    ) -> ET.Element:
        ref = f"{self._col_to_letters(col)}{row_idx}"
        for c in row_el.findall(self._q("c")):
            if c.attrib.get("r") == ref:
                return c
        # 创建新 cell，并尽量放到正确位置（按列顺序）
        attrib = {"r": ref}
        style = (self._style_template.get("col_style") or {}).get(col)
        if style is not None:
            attrib["s"] = style
        new_c = ET.Element(self._q("c"), attrib)

        inserted = False
        cells = list(row_el.findall(self._q("c")))
        for i, c in enumerate(cells):
            cref = c.attrib.get("r") or ""
            try:
                ccol, _ = self._cell_ref_to_col_row(cref)
            except Exception:
                continue
            if ccol > col:
                row_el.insert(i, new_c)
                inserted = True
                break
        if not inserted:
            row_el.append(new_c)
        return new_c

    def _set_cell_inline_str(self, c: ET.Element, value: str):
        # 保留样式索引 s，重置值节点
        for child in list(c):
            c.remove(child)
        c.attrib.pop("t", None)
        c.attrib["t"] = "inlineStr"
        is_el = ET.SubElement(c, self._q("is"))
        t_el = ET.SubElement(is_el, self._q("t"))
        t_el.text = value or ""
        if (t_el.text or "").strip() != (t_el.text or ""):
            t_el.attrib[self.XML_SPACE] = "preserve"

    def write_row_values(self, row_idx: int, values_by_col: dict[int, str]):
        if not values_by_col:
            return
        row_el = self._ensure_row_exists(row_idx)
        for col, value in values_by_col.items():
            target_row, target_col = self._redirect_merged_target(row_idx, int(col))
            target_row_el = self._ensure_row_exists(target_row)
            c = self._get_or_create_cell(target_row_el, target_row, target_col)
            self._set_cell_inline_str(c, str(value) if value is not None else "")

    def extend_table_and_filter_refs(self):
        if self._need_last_row <= 0:
            return
        self._extend_autofilter_ref()
        self._extend_tables_ref()

    def _extend_autofilter_ref(self):
        af = self._sheet_xml.getroot().find(self._q("autoFilter"))
        if af is None:
            return
        ref = af.attrib.get("ref")
        if not ref or ":" not in ref:
            return
        min_col, min_row, max_col, max_row = self._range_boundaries(ref)
        if self._need_last_row <= max_row:
            return
        af.attrib["ref"] = (
            f"{self._col_to_letters(min_col)}{min_row}:{self._col_to_letters(max_col)}{self._need_last_row}"
        )

    def _extend_tables_ref(self):
        # tableParts 在 sheet 中，真正的 ref 在 xl/tables/tableX.xml
        if self._sheet_rels_xml is None:
            return
        table_parts = self._sheet_xml.getroot().find(self._q("tableParts"))
        if table_parts is None:
            return
        for tp in table_parts.findall(self._q("tablePart")):
            rid = tp.attrib.get(self._qr("id"))
            if not rid:
                continue
            rel = self._sheet_rels_xml.getroot().find(
                f".//{{{self.NS_PKG_REL}}}Relationship[@Id='{rid}']"
            )
            if rel is None:
                continue
            target = rel.attrib.get("Target")
            if not target:
                continue
            # Target 形如 '../tables/table1.xml'
            table_path = posixpath.normpath(
                posixpath.join(posixpath.dirname(self._sheet_path), target)
            )
            if table_path.startswith("../"):
                table_path = posixpath.normpath(
                    posixpath.join(posixpath.dirname(self._sheet_path), table_path)
                )
            if table_path not in self._original_entries:
                # 兜底：有的 target 没带 ../
                table_path2 = posixpath.normpath(
                    posixpath.join("xl", target.lstrip("/"))
                )
                table_path = (
                    table_path2 if table_path2 in self._original_entries else table_path
                )
            if table_path not in self._original_entries:
                continue

            table_xml = self._parse_xml(self._original_entries[table_path])
            table_root = table_xml.getroot()
            tref = table_root.attrib.get("ref")
            if tref and ":" in tref:
                min_col, min_row, max_col, max_row = self._range_boundaries(tref)
                if (
                    self._need_last_row > max_row
                    and self._need_last_row >= self._data_start_row
                ):
                    table_root.attrib["ref"] = (
                        f"{self._col_to_letters(min_col)}{min_row}:{self._col_to_letters(max_col)}{self._need_last_row}"
                    )
            taf = table_root.find(self._q("autoFilter"))
            if taf is not None:
                aref = taf.attrib.get("ref")
                if aref and ":" in aref:
                    min_col, min_row, max_col, max_row = self._range_boundaries(aref)
                    if (
                        self._need_last_row > max_row
                        and self._need_last_row >= self._data_start_row
                    ):
                        taf.attrib["ref"] = (
                            f"{self._col_to_letters(min_col)}{min_row}:{self._col_to_letters(max_col)}{self._need_last_row}"
                        )

            self._original_entries[table_path] = self._serialize_xml(table_xml)

    def _serialize_xml(self, tree: ET.ElementTree) -> bytes:
        return ET.tostring(tree.getroot(), encoding="utf-8", xml_declaration=True)

    def _range_boundaries(self, ref: str) -> tuple[int, int, int, int]:
        a, b = ref.split(":")
        c1, r1 = self._cell_ref_to_col_row(a)
        c2, r2 = self._cell_ref_to_col_row(b)
        return min(c1, c2), min(r1, r2), max(c1, c2), max(r1, r2)

    def update_dimension(self):
        dim = self._sheet_xml.getroot().find(self._q("dimension"))
        if dim is None:
            return
        ref = dim.attrib.get("ref")
        if not ref or ":" not in ref:
            return
        min_col, min_row, max_col, max_row = self._range_boundaries(ref)
        new_max_row = max(max_row, self._need_last_row)
        if new_max_row == max_row:
            return
        dim.attrib["ref"] = (
            f"{self._col_to_letters(min_col)}{min_row}:{self._col_to_letters(max_col)}{new_max_row}"
        )

    def to_bytes(self) -> bytes:
        # 写回 sheet xml
        self._original_entries[self._sheet_path] = self._serialize_xml(self._sheet_xml)
        out = io.BytesIO()
        with zipfile.ZipFile(out, "w", compression=zipfile.ZIP_DEFLATED) as zf:
            for name, data in self._original_entries.items():
                zf.writestr(name, data)
        return out.getvalue()
